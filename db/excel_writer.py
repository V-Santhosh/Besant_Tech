import os
from openpyxl import Workbook, load_workbook
from config import EXCEL_FILE

def write_to_excel(labels, row_data):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "User Details"
        ws.append(labels)
        ws.append(row_data)
        wb.save(EXCEL_FILE)
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append(row_data)
        wb.save(EXCEL_FILE)